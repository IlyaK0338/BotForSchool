import logging
import os
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import Message, Update
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram import BaseMiddleware
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
import openpyxl
import shutil
import random
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import asyncio
from typing import Dict, Any, Callable, Awaitable
from telegram.ext import CommandHandler

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Инициализация бота
API_TOKEN = '7917610346:AAFyPpHk1UYx2xEUp4osqKl7TbIbyrXIKJI'
storage = MemoryStorage()
bot = Bot(
    token=API_TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=storage)


class ThrottlingMiddleware(BaseMiddleware):
    def __init__(self, limit: int = 5, interval: int = 1):
        self.rate_limit = limit
        self.interval = timedelta(seconds=interval)
        self.user_timestamps: Dict[int, list] = {}
        super().__init__()

    async def __call__(
            self,
            handler: Callable[[Update, Dict[str, Any]], Awaitable[Any]],
            event: Update,
            data: Dict[str, Any]
    ) -> Any:
        if not (message := getattr(event, 'message', None)) or not message.from_user:
            return await handler(event, data)

        user_id = message.from_user.id
        now = datetime.now()

        if user_id not in self.user_timestamps:
            self.user_timestamps[user_id] = []

        self.user_timestamps[user_id] = [
            t for t in self.user_timestamps[user_id]
            if now - t < self.interval
        ]

        if len(self.user_timestamps[user_id]) >= self.rate_limit:
            await message.answer("Слишком много запросов. Пожалуйста, подождите.")
            return

        self.user_timestamps[user_id].append(now)
        return await handler(event, data)


# Добавляем middleware
dp.update.outer_middleware(ThrottlingMiddleware())


# Состояния FSM
class Form(StatesGroup):
    waiting_for_parents = State()
    waiting_for_name = State()
    waiting_for_class = State()
    waiting_for_email = State()
    waiting_for_code = State()
    waiting_for_baby = State()
    waiting_for_code_user = State()
    waiting = State()


# Создание файла Excel
def create_new_excel_file():
    try:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'ФИО'
        ws['B1'] = 'Username'
        ws['C1'] = 'Роль'
        ws["D1"] = "Дети"
        ws["E1"] = "Код"
        ws["F1"] = "Статус регистрации"
        ws["G1"] = "ID"
        wb.save('users.xlsx')
        logger.info("Создан новый файл users.xlsx")
        return True
    except Exception as e:
        logger.error(f"Ошибка при создании файла Excel: {e}")
        return False
    finally:
        if 'wb' in locals():
            wb.close()


def repair_excel_file():
    try:
        if os.path.exists('users.xlsx'):
            backup_file = 'users_backup.xlsx'
            shutil.copy2('users.xlsx', backup_file)
            logger.info(f"Создана резервная копия: {backup_file}")

            try:
                wb = load_workbook('users.xlsx')
                wb.close()
                logger.info("Файл успешно открыт, повреждений не обнаружено")
                return True
            except:
                logger.warning("Файл повреждён, создаём новый")
                return create_new_excel_file()
        else:
            return create_new_excel_file()
    except Exception as e:
        logger.error(f"Ошибка при восстановлении файла: {e}")
        return False


# Инициализация файла Excel
if not repair_excel_file():
    logger.error("Не удалось инициализировать файл Excel")
    exit(1)


async def generate_code():
    return ''.join(random.choices('0123456789', k=6))


# Работа с Excel
async def safe_load_workbook(filename):
    try:
        return load_workbook(filename)
    except Exception as e:
        logger.error(f"Ошибка загрузки файла {filename}: {e}")
        if repair_excel_file():
            return load_workbook(filename)
        raise


async def add_to_db(fio, username, role, children, id, state: FSMContext):
    try:
        if await is_user_not_exist(username):
            wb = await safe_load_workbook('users.xlsx')
            try:
                ws = wb.active
                password = await generate_code()
                new_row = ws.max_row + 1
                await state.update_data(row=new_row)
                ws[f'A{new_row}'] = fio
                ws[f'B{new_row}'] = username
                ws[f'C{new_row}'] = role
                ws[f"D{new_row}"] = children
                ws[f"E{new_row}"] = password
                ws[f"F{new_row}"] = "В процессе регистрации"
                ws[f"G{new_row}"] = id
                await state.update_data(code=password)

                wb.save('users.xlsx')
                logger.info(f"Пользователь {username} добавлен в базу, код готов к выдаче")
                return True
            finally:
                wb.close()
        return False
    except Exception as e:
        logger.error(f"Ошибка в add_to_db: {e}")
        raise


async def end_reg(username, row):
    try:
        wb = await safe_load_workbook("users.xlsx")
        try:
            ws = wb.active
            ws[f"F{row}"] = "Регистрация завершена"
            wb.save('users.xlsx')
            logger.info(f"Пользователь {username} завершил регистрацию")
            return True
        except Exception as e:
            logger.error(f"Ошибка в end_reg: {e}")
            return False
        finally:
            wb.close()
    except Exception as e:
        logger.error(f"Произошла ошибка в {e}")
        return False


async def is_user_not_exist(username):
    try:
        wb = await safe_load_workbook('users.xlsx')
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] and str(row[1]).strip().lower() == username.strip().lower():
                    wb.close()
                    return False
            return True
        except Exception as e:
            logger.error(f"Ошибка при проверке пользователя: {e}")
            return False
        finally:
            wb.close()
    except Exception as e:
        logger.error(f"Ошибка в is_user_not_exist: {e}")
        return False


async def delete_user_by_username(username, file_path="users.xlsx"):
    try:
        wb = await safe_load_workbook(file_path)
        try:
            ws = wb.active
            rows_to_delete = []
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=2).value == username:
                    rows_to_delete.append(row)

            for row in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row)

            if rows_to_delete:
                wb.save(file_path)
                logger.info(f"Пользователь {username} удален из базы")
                return True
            return False
        finally:
            wb.close()
    except Exception as e:
        logger.error(f"Ошибка в delete_user_by_username: {e}")
        raise


async def get_user_reg_status(username):
    try:
        wb = await safe_load_workbook('users.xlsx')
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                if str(row[1].value).strip().lower() == username.strip().lower():
                    status = row[5].value if row[5].value else ""
                    return status
            return None
        finally:
            wb.close()
    except Exception as e:
        logger.error(f"Ошибка при получении статуса регистрации: {e}")
        return None


@dp.message(Command("start"))
async def cmd_start(message: Message):
    await message.answer(
        "Привет! Это бот-информатор по поступлению в 10 класс ГБОУ 1811\n"
        "Доступные команды:\n"
        "/start - старт бота\n"
        "/reg - регистрация\n")


@dp.message(Command("give_reg"))
async def give_reg(message: Message):
    username = message.from_user.username
    if username == "Ilya_k0338":
        wb = await safe_load_workbook('users.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            target_user_id = str(row[6].value).strip().lower()
            msg_text = "Это тестовое сообщение для пользователя!"
            await bot.send_message(
                chat_id=target_user_id,
                text=msg_text
            )
            await message.answer(f"✅ Сообщение отправлено пользователю (username: "
                                 f"{str(row[1].value).strip().lower()})")


@dp.message(Command("results"))
async def cmd_results(message: Message):
    await message.answer("Результатов пока что нет(")

@dp.message(Command("reg"))
async def cmd_reg(message: Message, state: FSMContext):
    username = message.from_user.username
    if not username:
        await message.answer("Для регистрации у вас должен быть username в Telegram!")
        return

    reg_status = await get_user_reg_status(username)

    if reg_status is not None:
        if "завершена" in reg_status:
            await message.answer("Вы уже завершили регистрацию!")
            return
        else:
            await delete_user_by_username(username)
            await message.answer("Обнаружена незавершенная регистрация. Начинаем заново.")

    keyboard = types.ReplyKeyboardMarkup(
        keyboard=[
            [types.KeyboardButton(text="Я родитель")],
            [types.KeyboardButton(text="Я ученик")]
        ],
        resize_keyboard=True
    )
    await state.set_state(Form.waiting_for_parents)
    await message.answer("Кто вы?", reply_markup=keyboard)


@dp.message(Command("cancel"))
async def cmd_cancel(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is None:
        await message.answer("Нет активных действий для отмены")
        return

    username = message.from_user.username
    if username:
        await delete_user_by_username(username)

    await state.clear()
    await message.answer("Действие отменено. Все данные удалены.")


# Обработчики состояний
@dp.message(Form.waiting_for_parents)
async def process_parents(message: Message, state: FSMContext):
    if message.text == "Я родитель":
        await state.update_data(role="parent")
        await message.answer("Введите ваше ФИО:", reply_markup=types.ReplyKeyboardRemove())
        await state.set_state(Form.waiting_for_name)
    elif message.text == "Я ученик":
        await state.update_data(role="student", children="-")
        await message.answer("Введите ваше ФИО:", reply_markup=types.ReplyKeyboardRemove())
        await state.set_state(Form.waiting_for_name)
    else:
        await message.answer("Пожалуйста, выберите один из предложенных вариантов")


@dp.message(Form.waiting_for_name)
async def process_name(message: Message, state: FSMContext):
    await state.update_data(fio=message.text)
    data = await state.get_data()
    username = message.from_user.username
    userid = message.from_user.id

    if data['role'] == "student":
        if await add_to_db(data["fio"], username, data['role'], data["children"], userid, state):
            await message.answer("Введите код, который вам скажет волонтер")
            await state.set_state(Form.waiting_for_code_user)
    else:
        await message.answer("Теперь введите ФИО своих детей, например: Иванов Иван Иванович, "
                             "если детей больше одного, то пишите их через запятую (ФИО1, ФИО2, ...)")
        await state.set_state(Form.waiting_for_baby)


@dp.message(Form.waiting_for_baby)
async def get_baby(message: Message, state: FSMContext):
    username = message.from_user.username
    userid = message.from_user.id
    await state.update_data(children=message.text)
    data = await state.get_data()
    if await add_to_db(data["fio"], username, data['role'], data["children"], userid, state):
        await message.answer("Теперь введите код, который вам скажет волонтер:")
        await state.set_state(Form.waiting_for_code_user)


@dp.message(Form.waiting_for_code_user)
async def waiting_code_user(message: Message, state: FSMContext):
    data = await state.get_data()
    username = message.from_user.username

    if 'code' not in data or 'row' not in data:
        await message.answer("Произошла ошибка. Пожалуйста, начните регистрацию заново.")
        await state.clear()
        return

    attempts = data.get('attempts', 0)

    if data['code'] == message.text:
        if await end_reg(username, data['row']):
            await message.answer("Вы успешно зарегистрировались!")
            await state.clear()
    else:
        attempts += 1
        await state.update_data(attempts=attempts)

        if attempts < 2:
            await message.answer(f"Неверный код. У вас осталась {2 - attempts} попытка. Попробуйте еще раз:")
        else:
            await message.answer("Вы исчерпали все попытки ввода кода. Пожалуйста, начните регистрацию заново.")
            await delete_user_by_username(username)
            await state.clear()


async def main():
    await dp.start_polling(bot)


if __name__ == '__main__':
    try:
        logger.info("Бот запущен")
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logger.info("Бот остановлен")
    except Exception as e:
        logger.error(f"Ошибка в основном цикле: {e}")
